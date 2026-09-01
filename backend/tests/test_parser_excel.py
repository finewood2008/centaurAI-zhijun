import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from parser import is_supported, parse_file


class ExcelParserTest(unittest.TestCase):
    def test_xlsx_extracts_sheet_names_rows_and_types(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sales.xlsx"
            workbook = Workbook()
            sales = workbook.active
            sales.title = "销售明细"
            sales.append(["客户", "金额", "签约日期", "已回款"])
            sales.append(["华东客户", 12500, date(2026, 7, 18), True])
            sales.append(["华南客户", None, datetime(2026, 7, 19, 9, 30), False])
            empty = workbook.create_sheet("空表")
            empty["C4"] = None
            notes = workbook.create_sheet("备注")
            notes.append(["说明", "包含\n换行"])
            workbook.save(path)
            workbook.close()

            parsed = parse_file(str(path))

            self.assertTrue(is_supported(str(path)))
            self.assertEqual(parsed["file_type"], "text")
            self.assertIn("工作表：销售明细", parsed["text"])
            self.assertIn("华东客户\t12500\t2026-07-18\tTRUE", parsed["text"])
            self.assertIn("华南客户\t\t2026-07-19 09:30:00\tFALSE", parsed["text"])
            self.assertIn("工作表：备注", parsed["text"])
            self.assertIn("包含 / 换行", parsed["text"])
            self.assertNotIn("工作表：空表", parsed["text"])


if __name__ == "__main__":
    unittest.main()
